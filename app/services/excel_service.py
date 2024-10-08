import logging
import unicodedata
import pandas as pd
from fastapi import HTTPException
from app.core.config import settings
from app.services.word_service import generate_word_document
import os
import openpyxl

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Fonction pour normaliser une chaîne de caractères
def normalize_string(s):
    if not isinstance(s, str):
        s = str(s)
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn').lower()


def normalize_name(name):
    """Normalise les noms pour comparaison."""
    if not name:
        return ""
    # Supprimer les accents
    name = unicodedata.normalize('NFD', name).encode('ascii', 'ignore').decode('utf-8')
    # Convertir en majuscules
    name = name.upper()
    # Supprimer les espaces en début et fin
    name = name.strip()
    return name

# Fonction pour traiter un fichier Excel
def process_excel_file(file_path: str, output_dir: str) -> list:
    try:
        # Load data from Excel file
        logger.info("Loading Excel file.")
        df_titles = pd.read_excel(file_path, header=None)
        df_students = pd.read_excel(file_path, header=1)
        
        # Rename columns for consistency
        df_students = df_students.rename(columns={
            'DatedeNaissance': 'Date de Naissance',
            'NomSite': 'Nom Site',
            'CodeGroupe': 'Code Groupe',
            'NomGroupe': 'Nom Groupe',
            'EtenduGroupe': 'Étendu Groupe',
            'ABSjustifiées': 'ABS justifiées',
            'ABSinjustifiées': 'ABS injustifiées',
        })
        logger.info(f"{len(df_students)} students found in the file.")
        
        # Define configurations for different cases
        cases = {
            "M1_S1": {
                "key": "M1_S1",
                "titles_row": df_titles.iloc[0, 2:22].tolist(),
                "template_word": settings.M1_S1_MAPI_TEMPLATE_WORD,
                "grade_column_indices": [3, 4, 5, 7, 9, 10, 12, 13, 14, 15, 16, 17, 19, 20, 21],
                "ects_sum_indices": {
                    'UE1': [1, 2, 3],
                    'UE2': [4],
                    'UE3': [5, 6],
                    'UE4': [7, 11],
                    'UE5': [13, 14, 15]
                },
                "hidden_ects": [8, 9, 10, 12]
            },
            "M1_S2": {
                "key": "M1_S2",
                "titles_row": df_titles.iloc[0, 2:22].tolist(),
                "template_word": settings.M1_S2_MAPI_TEMPLATE_WORD,
                "grade_column_indices": [3, 4, 5, 7, 8, 10, 11, 12, 13, 14, 15, 16, 18, 19, 20, 21],
                "ects_sum_indices": {
                    'UE1': [1, 2, 3],
                    'UE2': [4, 5],
                    'UE3': [6, 7, 8, 12],
                    'UE4': [13, 14, 15, 16],
                },
                "hidden_ects": [9, 10, 11]
            },
            "M2_S3_MAGI": {
                "key": "M2_S3_MAGI",
                "titles_row": df_titles.iloc[0, 2:19].tolist(),
                "template_word": settings.M2_S3_MAGI_TEMPLATE_WORD,
                "grade_column_indices": [3, 4, 6, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18],
                "ects_sum_indices": {
                    'UE1': [1, 2],
                    'UE2': [3],
                    'UE3': [4, 5, 6, 7, 8, 9],
                    'UE4': [10, 11, 12, 13],
                },
                "hidden_ects": [4, 8, 9]
            },
            "M2_S3_MEFIM": {
                "key": "M2_S3_MEFIM",
                "titles_row": df_titles.iloc[0, 2:19].tolist(),
                "template_word": settings.M2_S3_MAGI_TEMPLATE_WORD,
                "grade_column_indices": [3, 4, 6, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18],
                "ects_sum_indices": {
                    'UE1': [1, 2],
                    'UE2': [3],
                    'UE3': [4, 5, 6, 7, 8, 9],
                    'UE4': [10, 11, 12, 13],
                },
                "hidden_ects": [4, 8, 9]
            },
            "M2_S3_MAPI": {
                "key": "M2_S3_MAPI",
                "titles_row": df_titles.iloc[0, 2:20].tolist(),
                "template_word": settings.M2_S3_MAPI_TEMPLATE_WORD,
                "grade_column_indices": [3, 4, 6, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19],
                "ects_sum_indices": {
                    'UE1': [1, 2],
                    'UE2': [3],
                    'UE3': [4, 5, 6, 7, 8, 9],
                    'UE4': [10, 11, 12, 13, 14],
                },
                "hidden_ects": [4, 8, 9]
            },
            "M2_S4": {
                "key": "M2_S4",
                "titles_row": df_titles.iloc[0, 2:17].tolist(),
                "template_word": settings.M2_S4_MAPI_TEMPLATE_WORD,
                "grade_column_indices": [3, 5, 6, 8, 9, 10, 11, 12, 14, 15, 16],
                "ects_sum_indices": {
                    'UE1': [1],
                    'UE2': [2, 3],
                    'UE3': [4, 5, 8],
                    'UE4': [9, 10, 11],
                },
                "hidden_ects": [6, 7]
            }
        }
        
        # Check if the file exists
        if not os.path.exists(file_path):
            raise HTTPException(status_code=400, detail="File not found")

        # Determine the case key based on file name comparison
        filename = os.path.basename(file_path)
        case_key = None
        for key, templates in settings.TEMPLATE_MAPPING.items():
            if filename in [os.path.basename(template) for template in templates]:
                case_key = key
                break
        
        if case_key is None:
            raise HTTPException(status_code=400, detail="Unknown Excel template")

        case_config = cases[case_key]

        # List to store paths of generated bulletins
        bulletin_paths = []
        for index, student_data in df_students.iterrows():
            # Ensure all fields are strings to avoid issues with normalize_string
            student_data = student_data.fillna('').astype(str)
            
            # Check if essential fields are empty to skip empty bulletins
            if not student_data["Nom"] or not student_data["CodeApprenant"]:
                logger.info(f"Skipping empty bulletin for row {index} with data: {student_data}")
                continue

            # Generate Word document for the student
            bulletin_path = generate_word_document(student_data, case_config, case_config["template_word"], output_dir)
            if bulletin_path:
                bulletin_paths.append(bulletin_path)
                logger.info(f"Bulletin generated for {student_data.get('Nom', 'N/A')}: {bulletin_path}")

        return bulletin_paths
    except Exception as e:
        # Log in case of error during Excel file processing
        logger.error("Error processing Excel file", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")

def extract_appreciations_from_word(word_file_path):
    import unicodedata
    import docx
    doc = docx.Document(word_file_path)
    appreciations = {}
    
    # Fonction de normalisation des chaînes de caractères
    def normalize_string(s):
        return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn').upper().strip()
    
    for para in doc.paragraphs:
        if para.text:
            parts = para.text.split(':')
            if len(parts) == 2:
                name, appreciation = parts
                normalized_name = normalize_string(name)
                appreciations[normalized_name] = appreciation.strip()
    return appreciations

def update_excel_with_appreciations(template_wb, appreciations, columns_config):
    template_ws = template_wb.active
    appreciation_column_index = columns_config.get('appreciation_column_index_template', 31)  # Colonne par défaut AE

    for row in range(2, template_ws.max_row + 1):
        student_name = template_ws.cell(row=row, column=columns_config['name_column_index_template']).value
        if student_name:
            normalized_student_name = normalize_name(student_name)
            
            # Rechercher l'appréciation normalisée dans le dictionnaire
            for key, appreciation in appreciations.items():
                if normalize_name(key) == normalized_student_name:
                    template_ws.cell(row=row, column=appreciation_column_index).value = appreciation
                    break
            else:
                print(f"Appreciation non trouvée pour: {student_name}")

    return template_wb